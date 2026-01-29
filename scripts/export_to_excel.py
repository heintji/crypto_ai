import os
import csv
from datetime import datetime
from collections import defaultdict

LOG_CSV = os.path.join("logs", "paper_trades.csv")
OUT_XLSX = os.path.join("logs", "reports", "trading_dashboard.xlsx")

def try_import_openpyxl():
    try:
        import openpyxl  # noqa
        return True
    except Exception:
        return False

def parse_dt(s: str):
    try:
        # verwacht: 2026-01-28T12:34:56
        return datetime.fromisoformat((s or "").strip())
    except Exception:
        return None

def safe_float(x, default=0.0):
    try:
        return float(x)
    except Exception:
        return default

def read_trades():
    if not os.path.isfile(LOG_CSV):
        raise FileNotFoundError(f"Bestand niet gevonden: {LOG_CSV}")

    rows = []
    with open(LOG_CSV, "r", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            dt = parse_dt(row.get("datetime", ""))
            sym = (row.get("symbol") or "").strip().upper()
            side = (row.get("side") or "").strip().upper()
            price = safe_float(row.get("price"), 0.0)
            qty = safe_float(row.get("qty"), 0.0)
            balance = safe_float(row.get("balance"), 0.0)
            pnl = safe_float(row.get("pnl"), 0.0)
            meta = (row.get("meta") or "").strip()

            rows.append({
                "datetime": dt,
                "datetime_raw": row.get("datetime", ""),
                "symbol": sym,
                "side": side,
                "price": price,
                "qty": qty,
                "balance": balance,
                "pnl": pnl,
                "meta": meta
            })

    # sort op tijd
    rows.sort(key=lambda x: x["datetime"] or datetime.min)
    return rows

def iso_week_key(dt: datetime):
    if not dt:
        return "UNKNOWN"
    y, w, _ = dt.isocalendar()
    return f"{y}-W{w:02d}"

def build_aggregates(rows):
    sells = [x for x in rows if x["side"] == "SELL"]

    by_coin = defaultdict(lambda: {"closed": 0, "wins": 0, "losses": 0, "breakeven": 0, "pnl": 0.0})
    by_week = defaultdict(lambda: {"closed": 0, "wins": 0, "losses": 0, "breakeven": 0, "pnl": 0.0})

    for s in sells:
        sym = s["symbol"]
        wk = iso_week_key(s["datetime"])
        p = s["pnl"]

        by_coin[sym]["closed"] += 1
        by_week[wk]["closed"] += 1

        if p > 0:
            by_coin[sym]["wins"] += 1
            by_week[wk]["wins"] += 1
        elif p < 0:
            by_coin[sym]["losses"] += 1
            by_week[wk]["losses"] += 1
        else:
            by_coin[sym]["breakeven"] += 1
            by_week[wk]["breakeven"] += 1

        by_coin[sym]["pnl"] += p
        by_week[wk]["pnl"] += p

    # equity curve (balance over time)
    equity = []
    for r in rows:
        if r["datetime"]:
            equity.append((r["datetime"], r["balance"]))

    return sells, by_coin, by_week, equity

def ensure_sheet(wb, name):
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
        return ws
    return wb.create_sheet(title=name)

def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

def write_excel(rows, sells, by_coin, by_week, equity):
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.chart import LineChart, BarChart, Reference

    if os.path.isfile(OUT_XLSX):
        wb = openpyxl.load_workbook(OUT_XLSX)
    else:
        wb = openpyxl.Workbook()

    # Maak eerste default sheet netjes
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        ws0 = wb["Sheet"]
        ws0.title = "README"
        ws0["A1"] = "Crypto_AI Trading Dashboard"
        ws0["A2"] = "Dit bestand wordt automatisch gevuld vanuit logs/paper_trades.csv"
        ws0["A1"].font = Font(bold=True)

    ws_trades = ensure_sheet(wb, "Trades")
    ws_week = ensure_sheet(wb, "Weekly")
    ws_coin = ensure_sheet(wb, "PerCoin")
    ws_equity = ensure_sheet(wb, "EquityCurve")

    # ============ Trades ============
    headers = ["datetime", "symbol", "side", "price", "qty", "balance", "pnl", "meta"]
    ws_trades.append(headers)
    ws_trades[1][0].font = Font(bold=True)
    for hcell in ws_trades[1]:
        hcell.font = Font(bold=True)

    for r in rows:
        dt_str = r["datetime"].isoformat(sep=" ", timespec="seconds") if r["datetime"] else r["datetime_raw"]
        ws_trades.append([dt_str, r["symbol"], r["side"], r["price"], r["qty"], r["balance"], r["pnl"], r["meta"]])

    autosize(ws_trades)

    # ============ Weekly ============
    ws_week.append(["week", "closed", "wins", "losses", "breakeven", "winrate_%", "pnl_eur"])
    for hcell in ws_week[1]:
        hcell.font = Font(bold=True)

    week_keys = sorted(by_week.keys())
    for wk in week_keys:
        s = by_week[wk]
        closed = s["closed"]
        winrate = (s["wins"] / closed) * 100 if closed else 0.0
        ws_week.append([wk, closed, s["wins"], s["losses"], s["breakeven"], round(winrate, 2), round(s["pnl"], 2)])

    autosize(ws_week)

    # ============ PerCoin ============
    ws_coin.append(["symbol", "closed", "wins", "losses", "breakeven", "winrate_%", "pnl_eur"])
    for hcell in ws_coin[1]:
        hcell.font = Font(bold=True)

    coin_keys = sorted(by_coin.keys(), key=lambda k: by_coin[k]["pnl"], reverse=True)
    for sym in coin_keys:
        s = by_coin[sym]
        closed = s["closed"]
        winrate = (s["wins"] / closed) * 100 if closed else 0.0
        ws_coin.append([sym, closed, s["wins"], s["losses"], s["breakeven"], round(winrate, 2), round(s["pnl"], 2)])

    autosize(ws_coin)

    # ============ Equity Curve ============
    ws_equity.append(["datetime", "balance"])
    for hcell in ws_equity[1]:
        hcell.font = Font(bold=True)

    for dt, bal in equity:
        ws_equity.append([dt.isoformat(sep=" ", timespec="seconds"), round(bal, 2)])

    autosize(ws_equity)

    # ============ Charts (in Weekly sheet) ============
    # Maak ruimte voor charts rechts
    chart_anchor_col = 9  # column I
    chart_anchor_row = 2

    # 1) PnL per week (bar)
    if ws_week.max_row >= 2:
        bar = BarChart()
        bar.title = "PnL per week (€)"
        data = Reference(ws_week, min_col=7, min_row=1, max_row=ws_week.max_row)  # pnl
        cats = Reference(ws_week, min_col=1, min_row=2, max_row=ws_week.max_row)  # week labels
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 10
        bar.width = 18
        ws_week.add_chart(bar, f"I2")

    # 2) Balance over tijd (line) op Equity sheet
    if ws_equity.max_row >= 2:
        line = LineChart()
        line.title = "Balance (equity) over tijd"
        data = Reference(ws_equity, min_col=2, min_row=1, max_row=ws_equity.max_row)
        cats = Reference(ws_equity, min_col=1, min_row=2, max_row=ws_equity.max_row)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.height = 10
        line.width = 28
        ws_equity.add_chart(line, "D2")

    # 3) PnL per coin (bar) op PerCoin sheet
    if ws_coin.max_row >= 2:
        bar2 = BarChart()
        bar2.title = "PnL per coin (€)"
        data2 = Reference(ws_coin, min_col=7, min_row=1, max_row=ws_coin.max_row)
        cats2 = Reference(ws_coin, min_col=1, min_row=2, max_row=ws_coin.max_row)
        bar2.add_data(data2, titles_from_data=True)
        bar2.set_categories(cats2)
        bar2.height = 10
        bar2.width = 18
        ws_coin.add_chart(bar2, "I2")

    wb.save(OUT_XLSX)

def main():
    if not try_import_openpyxl():
        print("❌ openpyxl ontbreekt.")
        print("✅ Run dit eerst:")
        print("   pip install openpyxl")
        return

    rows = read_trades()
    sells, by_coin, by_week, equity = build_aggregates(rows)
    write_excel(rows, sells, by_coin, by_week, equity)

    total_sells = len(sells)
    total_pnl = sum(x["pnl"] for x in sells)
    print("✅ Excel dashboard gemaakt/geüpdatet:")
    print(f"- {os.path.abspath(OUT_XLSX)}")
    print(f"SELL trades: {total_sells} | Totaal PnL: €{total_pnl:.2f}")

if __name__ == "__main__":
    main()
